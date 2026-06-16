"""
v1.2 Data Preparation test suite.

Tests for read_uploaded_file_raw, get_raw_preview, and prepare_dataframe_from_raw
covering the 12 scenarios specified in the v1.2 requirements.
"""
import io

import pandas as pd
import pytest
from openpyxl import load_workbook

from src.io_utils import (
    get_raw_preview,
    prepare_dataframe_from_raw,
    read_uploaded_file_raw,
)
from src.delta_engine import run_delta
from src.reporting import export_to_excel


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

class _FakeFile:
    def __init__(self, content: bytes, name: str):
        self._buf = io.BytesIO(content)
        self.name = name

    def seek(self, pos: int, whence: int = 0) -> int:
        return self._buf.seek(pos, whence)

    def seekable(self) -> bool:
        return True

    def read(self, n: int = -1) -> bytes:
        return self._buf.read(n)

    def tell(self) -> int:
        return self._buf.tell()


def _csv_file(rows: list[list], name: str = "data.csv") -> _FakeFile:
    text = "\n".join(",".join(str(c) for c in r) for r in rows)
    return _FakeFile(text.encode("utf-8"), name)


def _make_raw(*rows) -> pd.DataFrame:
    df = pd.DataFrame([list(r) for r in rows], dtype=str)
    df.columns = range(len(df.columns))
    return df


# ---------------------------------------------------------------------------
# Test 1: Default header behavior matches v1.1 read_uploaded_file behavior
# ---------------------------------------------------------------------------

class TestDefaultHeaderBehavior:
    def test_row1_becomes_header_by_default(self):
        raw = _make_raw(
            ["CaseID", "Status", "Amount"],
            ["1001",   "Open",   "500"],
            ["1002",   "Closed", "750"],
        )
        df, meta, warns = prepare_dataframe_from_raw(raw)
        assert list(df.columns) == ["CaseID", "Status", "Amount"]
        assert len(df) == 2
        assert warns == []
        assert meta["rows_dropped_above_header"] == 0

    def test_clean_file_gives_zero_dropped_counts(self):
        raw = _make_raw(
            ["id", "val"],
            ["A",  "1"],
            ["B",  "2"],
        )
        _, meta, _ = prepare_dataframe_from_raw(raw)
        assert meta["rows_dropped_above_header"] == 0
        assert meta["rows_dropped_blank"] == 0
        assert meta["rows_dropped_after_end_row"] == 0
        assert meta["blank_headers_renamed"] == 0
        assert meta["duplicate_headers_renamed"] == 0


# ---------------------------------------------------------------------------
# Test 2: Header row 4 selected (rows 1-3 are title/notes)
# ---------------------------------------------------------------------------

class TestHeaderRowBelowRow1:
    def _messy_raw(self):
        return _make_raw(
            ["Monthly Reconciliation Report", "", ""],
            ["Exported from system: 2025-06-01", "", ""],
            ["", "", ""],
            ["CaseID", "Status", "Amount"],
            ["1001",   "Open",   "500"],
            ["1002",   "Closed", "750"],
        )

    def test_header_assigned_from_row_4(self):
        df, meta, _ = prepare_dataframe_from_raw(self._messy_raw(), header_row_index=3)
        assert list(df.columns) == ["CaseID", "Status", "Amount"]

    def test_rows_above_header_dropped(self):
        df, meta, _ = prepare_dataframe_from_raw(self._messy_raw(), header_row_index=3)
        assert meta["rows_dropped_above_header"] == 3
        assert len(df) == 2

    def test_data_rows_correct(self):
        df, _, _ = prepare_dataframe_from_raw(self._messy_raw(), header_row_index=3)
        assert df.iloc[0]["CaseID"] == "1001"
        assert df.iloc[1]["CaseID"] == "1002"


# ---------------------------------------------------------------------------
# Test 3: drop_rows_above=False preserves rows above header
# ---------------------------------------------------------------------------

class TestDropRowsAboveFalse:
    def test_rows_above_preserved_as_data(self):
        raw = _make_raw(
            ["agency_meta", "DoD"],
            ["CaseID",      "Status"],
            ["1001",        "Open"],
        )
        df, meta, _ = prepare_dataframe_from_raw(raw, header_row_index=1, drop_rows_above=False)
        assert meta["rows_dropped_above_header"] == 0
        assert len(df) == 2  # agency_meta row + data row

    def test_header_row_itself_excluded(self):
        raw = _make_raw(
            ["pre", "data"],
            ["CaseID", "Status"],
            ["1001",   "Open"],
        )
        df, _, _ = prepare_dataframe_from_raw(raw, header_row_index=1, drop_rows_above=False)
        # "CaseID" should be the column name, not a data value
        assert "CaseID" not in df["CaseID"].values
        assert "1001" in df["CaseID"].values

    def test_column_names_still_from_header_row(self):
        raw = _make_raw(
            ["notes", ""],
            ["id",    "val"],
            ["A",     "1"],
        )
        df, _, _ = prepare_dataframe_from_raw(raw, header_row_index=1, drop_rows_above=False)
        assert "id" in df.columns
        assert "val" in df.columns


# ---------------------------------------------------------------------------
# Test 4: Blank row removal
# ---------------------------------------------------------------------------

class TestDropBlankRows:
    def test_blank_rows_counted_in_metadata(self):
        raw = _make_raw(
            ["id"],
            ["1"],
            ["", ""],
            ["2"],
            ["",   ""],
        )
        df, meta, _ = prepare_dataframe_from_raw(raw, drop_blank_rows=True)
        assert meta["rows_dropped_blank"] == 2
        assert len(df) == 2

    def test_blank_rows_retained_when_disabled(self):
        raw = _make_raw(
            ["id"],
            ["1"],
            [""],
        )
        df, meta, _ = prepare_dataframe_from_raw(raw, drop_blank_rows=False)
        assert meta["rows_dropped_blank"] == 0
        assert len(df) == 2


# ---------------------------------------------------------------------------
# Test 5: End row filtering removes footer rows
# ---------------------------------------------------------------------------

class TestEndRowFiltering:
    def test_footer_row_excluded(self):
        raw = _make_raw(
            ["CaseID", "Amount"],
            ["1001",   "500"],
            ["1002",   "750"],
            ["Total",  "1250"],
        )
        df, meta, _ = prepare_dataframe_from_raw(raw, end_row_index=2)
        assert len(df) == 2
        assert "Total" not in df["CaseID"].values
        assert meta["rows_dropped_after_end_row"] == 1

    def test_end_row_metadata_records_selection(self):
        raw = _make_raw(["h"], ["1"], ["2"], ["Total"])
        _, meta, _ = prepare_dataframe_from_raw(raw, end_row_index=2)
        assert meta["end_row_selected"] == 2

    def test_no_end_row_metadata_is_none(self):
        raw = _make_raw(["h"], ["1"])
        _, meta, _ = prepare_dataframe_from_raw(raw)
        assert meta["end_row_selected"] is None
        assert meta["rows_dropped_after_end_row"] == 0


# ---------------------------------------------------------------------------
# Test 6: Blank headers renamed Unnamed_N
# ---------------------------------------------------------------------------

class TestBlankHeadersRenamed:
    def test_blank_header_gets_unnamed_label(self):
        raw = _make_raw(
            ["CaseID", "", "Status"],
            ["1001",   "x",  "Open"],
        )
        df, meta, warns = prepare_dataframe_from_raw(raw)
        assert "Unnamed_1" in df.columns
        assert meta["blank_headers_renamed"] == 1
        assert any("blank header" in w.lower() for w in warns)

    def test_multiple_blank_headers_get_sequential_labels(self):
        raw = _make_raw(
            ["id", "", ""],
            ["1",  "a", "b"],
        )
        df, meta, _ = prepare_dataframe_from_raw(raw)
        assert "Unnamed_1" in df.columns
        assert "Unnamed_2" in df.columns
        assert meta["blank_headers_renamed"] == 2


# ---------------------------------------------------------------------------
# Test 7: Duplicate headers renamed with _2, _3 suffix
# ---------------------------------------------------------------------------

class TestDuplicateHeadersRenamed:
    def test_duplicate_gets_numbered_suffix(self):
        raw = _make_raw(
            ["CaseID", "Amount", "Amount"],
            ["1001",   "500",    "600"],
        )
        df, meta, warns = prepare_dataframe_from_raw(raw)
        assert "Amount" in df.columns
        assert "Amount_2" in df.columns
        assert meta["duplicate_headers_renamed"] == 1
        assert any("Duplicate" in w for w in warns)

    def test_three_duplicate_headers(self):
        raw = _make_raw(
            ["Val", "Val", "Val"],
            ["a",   "b",   "c"],
        )
        df, meta, _ = prepare_dataframe_from_raw(raw)
        assert "Val" in df.columns
        assert "Val_2" in df.columns
        assert "Val_3" in df.columns
        assert meta["duplicate_headers_renamed"] == 2


# ---------------------------------------------------------------------------
# Test 8: Header row out of range
# ---------------------------------------------------------------------------

class TestHeaderRowOutOfRange:
    def test_raises_value_error_for_negative_index(self):
        raw = _make_raw(["id"], ["1"])
        with pytest.raises(ValueError, match="out of range"):
            prepare_dataframe_from_raw(raw, header_row_index=-1)

    def test_raises_value_error_for_index_beyond_file(self):
        raw = _make_raw(["id"], ["1"])
        with pytest.raises(ValueError, match="out of range"):
            prepare_dataframe_from_raw(raw, header_row_index=5)

    def test_empty_dataframe_raises(self):
        with pytest.raises(ValueError, match="empty"):
            prepare_dataframe_from_raw(pd.DataFrame())


# ---------------------------------------------------------------------------
# Test 9: End row at or before header row
# ---------------------------------------------------------------------------

class TestEndRowBeforeHeader:
    def test_end_row_before_header_ignored_with_warning(self):
        raw = _make_raw(["skip"], ["id"], ["1"], ["2"])
        df, meta, warns = prepare_dataframe_from_raw(
            raw, header_row_index=1, end_row_index=0
        )
        assert any("ignored" in w.lower() for w in warns)
        # Without the end row, both data rows should be present
        assert len(df) == 2

    def test_end_row_equal_to_header_ignored_with_warning(self):
        raw = _make_raw(["id"], ["1"])
        df, _, warns = prepare_dataframe_from_raw(
            raw, header_row_index=0, end_row_index=0
        )
        assert any("ignored" in w.lower() for w in warns)

    def test_end_row_exceeds_file_size_ignored_with_warning(self):
        raw = _make_raw(["id"], ["1"], ["2"])
        df, _, warns = prepare_dataframe_from_raw(raw, end_row_index=99)
        assert any("exceeds" in w.lower() for w in warns)
        assert len(df) == 2


# ---------------------------------------------------------------------------
# Test 10: Raw CSV read — no header assumed
# ---------------------------------------------------------------------------

class TestRawCsvRead:
    def test_first_row_treated_as_data(self):
        f = _csv_file([["CaseID", "Status"], ["1001", "Open"]])
        df = read_uploaded_file_raw(f)
        assert df.iloc[0, 0] == "CaseID"

    def test_columns_are_integer_indexed(self):
        f = _csv_file([["a", "b", "c"]])
        df = read_uploaded_file_raw(f)
        assert list(df.columns) == [0, 1, 2]

    def test_all_cell_values_are_strings(self):
        f = _csv_file([["id", "val"], ["1", "3.14"]])
        df = read_uploaded_file_raw(f)
        for col in df.columns:
            assert all(isinstance(v, str) for v in df[col])

    def test_seek_position_restored(self):
        f = _csv_file([["id"], ["1"]])
        read_uploaded_file_raw(f)
        assert f._buf.tell() == 0


# ---------------------------------------------------------------------------
# Test 11: Raw Excel read with selected sheet
# ---------------------------------------------------------------------------

class TestRawExcelRead:
    def _make_excel_bytes(self, rows: list[list]) -> bytes:
        df = pd.DataFrame(rows)
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as w:
            df.to_excel(w, sheet_name="Data", index=False, header=False)
        return buf.getvalue()

    def test_excel_read_raw_has_integer_columns(self):
        content = self._make_excel_bytes([["CaseID", "Status"], ["1001", "Open"]])
        f = _FakeFile(content, "data.xlsx")
        df = read_uploaded_file_raw(f, sheet_name="Data")
        assert list(df.columns) == [0, 1]

    def test_excel_first_row_is_data(self):
        content = self._make_excel_bytes([["CaseID", "Status"], ["1001", "Open"]])
        f = _FakeFile(content, "data.xlsx")
        df = read_uploaded_file_raw(f, sheet_name="Data")
        assert df.iloc[0, 0] == "CaseID"

    def test_excel_row_count_correct(self):
        content = self._make_excel_bytes([
            ["hdr1", "hdr2"],
            ["r1a",  "r1b"],
            ["r2a",  "r2b"],
        ])
        f = _FakeFile(content, "data.xlsx")
        df = read_uploaded_file_raw(f)
        assert len(df) == 3


# ---------------------------------------------------------------------------
# Test 12: get_raw_preview
# ---------------------------------------------------------------------------

class TestGetRawPreview:
    def _raw(self, n_rows: int = 30, n_cols: int = 5) -> pd.DataFrame:
        data = {i: [f"r{r}c{i}" for r in range(n_rows)] for i in range(n_cols)}
        df = pd.DataFrame(data)
        df.columns = range(n_cols)
        return df

    def test_source_row_column_present(self):
        df = get_raw_preview(self._raw())
        assert "Source Row" in df.columns

    def test_source_row_numbers_are_one_based(self):
        df = get_raw_preview(self._raw(n_rows=5))
        assert list(df["Source Row"]) == [1, 2, 3, 4, 5]

    def test_max_rows_respected(self):
        df = get_raw_preview(self._raw(n_rows=50), max_rows=25)
        assert len(df) == 25

    def test_max_cols_respected(self):
        raw = self._raw(n_cols=20)
        df = get_raw_preview(raw, max_cols=15)
        # 15 data cols + "Source Row"
        assert len(df.columns) == 16

    def test_does_not_mutate_original(self):
        raw = self._raw()
        original_cols = list(raw.columns)
        get_raw_preview(raw)
        assert list(raw.columns) == original_cols
        assert "Source Row" not in raw.columns

    def test_small_file_not_padded(self):
        raw = self._raw(n_rows=5)
        df = get_raw_preview(raw, max_rows=25)
        assert len(df) == 5


# ---------------------------------------------------------------------------
# Test 13: Preparation metadata appears in the Excel Analysis Metadata tab
# ---------------------------------------------------------------------------

class TestPreparationMetadataInExcel:
    def _run(self, df_a=None, df_b=None):
        if df_a is None:
            df_a = pd.DataFrame({"id": ["1", "2"], "val": ["a", "b"]})
        if df_b is None:
            df_b = pd.DataFrame({"id": ["1", "2"], "val": ["a", "c"]})
        return run_delta(df_a, df_b, ["id"], ["id"], ["val"], ["val"])

    def _params(self, wb) -> list[str]:
        ws = wb["Analysis Metadata"]
        return [str(ws.cell(row=r, column=1).value) for r in range(2, ws.max_row + 1)]

    def _values(self, wb) -> list[str]:
        ws = wb["Analysis Metadata"]
        return [str(ws.cell(row=r, column=2).value) for r in range(2, ws.max_row + 1)]

    def test_prep_rows_absent_when_no_metadata_provided(self):
        wb = load_workbook(io.BytesIO(export_to_excel(self._run())))
        params = self._params(wb)
        assert not any("Header Row Selected" in p for p in params)

    def test_source_prep_rows_present_when_metadata_provided(self):
        meta = {
            "header_row_selected": 3, "drop_rows_above": True, "drop_blank_rows": True,
            "end_row_selected": None, "rows_in_raw": 100, "rows_dropped_above_header": 3,
            "rows_dropped_blank": 0, "rows_dropped_after_end_row": 0,
            "rows_in_prepared": 96, "columns_in_prepared": 5,
            "blank_headers_renamed": 0, "duplicate_headers_renamed": 0,
        }
        wb = load_workbook(io.BytesIO(export_to_excel(self._run(), source_prep_metadata=meta)))
        params = self._params(wb)
        assert any("Source Header Row Selected" in p for p in params)
        assert any("Source Rows Dropped Above Header" in p for p in params)
        assert any("Source Prepared Row Count" in p for p in params)
        assert any("Source Blank Headers Renamed" in p for p in params)
        assert any("Source Duplicate Headers Renamed" in p for p in params)

    def test_comparison_prep_rows_present(self):
        meta = {
            "header_row_selected": 0, "drop_rows_above": True, "drop_blank_rows": True,
            "end_row_selected": 50, "rows_in_raw": 55, "rows_dropped_above_header": 0,
            "rows_dropped_blank": 2, "rows_dropped_after_end_row": 4,
            "rows_in_prepared": 48, "columns_in_prepared": 4,
            "blank_headers_renamed": 1, "duplicate_headers_renamed": 2,
        }
        wb = load_workbook(io.BytesIO(export_to_excel(self._run(), comparison_prep_metadata=meta)))
        params = self._params(wb)
        assert any("Comparison Header Row Selected" in p for p in params)
        assert any("Comparison Rows Dropped After End Row" in p for p in params)

    def test_header_row_displayed_one_based(self):
        meta = {
            "header_row_selected": 3, "drop_rows_above": True, "drop_blank_rows": True,
            "end_row_selected": None, "rows_in_raw": 10, "rows_dropped_above_header": 3,
            "rows_dropped_blank": 0, "rows_dropped_after_end_row": 0,
            "rows_in_prepared": 6, "columns_in_prepared": 3,
            "blank_headers_renamed": 0, "duplicate_headers_renamed": 0,
        }
        wb = load_workbook(io.BytesIO(export_to_excel(self._run(), source_prep_metadata=meta)))
        params = self._params(wb)
        values = self._values(wb)
        idx = next(i for i, p in enumerate(params) if "Source Header Row Selected" in p)
        assert values[idx] == "4"  # 3 (0-based) + 1 = 4

    def test_end_row_not_applied_shows_not_applied(self):
        meta = {
            "header_row_selected": 0, "drop_rows_above": True, "drop_blank_rows": True,
            "end_row_selected": None, "rows_in_raw": 5, "rows_dropped_above_header": 0,
            "rows_dropped_blank": 0, "rows_dropped_after_end_row": 0,
            "rows_in_prepared": 4, "columns_in_prepared": 2,
            "blank_headers_renamed": 0, "duplicate_headers_renamed": 0,
        }
        wb = load_workbook(io.BytesIO(export_to_excel(self._run(), source_prep_metadata=meta)))
        params = self._params(wb)
        values = self._values(wb)
        idx = next(i for i, p in enumerate(params) if "Source End Row Selected" in p)
        assert values[idx] == "Not applied"
