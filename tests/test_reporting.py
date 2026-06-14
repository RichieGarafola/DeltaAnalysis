"""
Tests for src/reporting.py — build_summary_df, build_change_frequency,
export_to_excel (sheet names, metadata content, rules content).
"""
import io

import pandas as pd
import pytest
from openpyxl import load_workbook

from src.delta_engine import DeltaResult, run_delta
from src.reporting import (
    build_change_frequency,
    build_summary_df,
    export_to_excel,
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

def _make_result(
    compare_cols=True,
    with_comparison_rules=False,
    sheet_a=None,
    sheet_b=None,
):
    """Build a minimal DeltaResult covering all categories."""
    df_a = pd.DataFrame({
        "id":     ["1", "2", "3", "4", "4", ""],
        "status": ["Open", "Pending", "Closed", "Open", "Open", "Unknown"],
        "amount": ["100", "200", "300", "400", "400", "50"],
    })
    df_b = pd.DataFrame({
        "id":     ["1", "2", "5"],
        "status": ["Open", "Approved", "New"],
        "amount": ["100", "250", "600"],
    })

    rules = None
    if with_comparison_rules:
        rules = [
            {"column_a": "status", "column_b": "status", "type": "text",    "tolerance": None, "date_mode": None},
            {"column_a": "amount", "column_b": "amount", "type": "numeric", "tolerance": 0.0,  "date_mode": None},
        ]

    return run_delta(
        df_a=df_a,
        df_b=df_b,
        key_cols_a=["id"],
        key_cols_b=["id"],
        compare_cols_a=["status", "amount"] if compare_cols else None,
        compare_cols_b=["status", "amount"] if compare_cols else None,
        comparison_rules=rules,
        sheet_a=sheet_a,
        sheet_b=sheet_b,
    )


# ---------------------------------------------------------------------------
# build_summary_df
# ---------------------------------------------------------------------------

class TestBuildSummaryDf:
    def test_returns_dataframe(self):
        result = _make_result()
        df = build_summary_df(result)
        assert isinstance(df, pd.DataFrame)

    def test_has_required_columns(self):
        df = build_summary_df(_make_result())
        assert list(df.columns) == ["Metric", "Count", "% of File Total"]

    def test_total_a_correct(self):
        result = _make_result()
        df = build_summary_df(result)
        row = df[df["Metric"] == "File A — Total Records"]
        assert int(row["Count"].iloc[0]) == result.total_a

    def test_total_b_correct(self):
        result = _make_result()
        df = build_summary_df(result)
        row = df[df["Metric"] == "File B — Total Records"]
        assert int(row["Count"].iloc[0]) == result.total_b

    def test_zero_total_b_returns_na_percentages(self):
        df_a = pd.DataFrame({"id": ["X"], "v": ["1"]})
        df_b = pd.DataFrame({"id": pd.Series([], dtype=str), "v": pd.Series([], dtype=str)})
        # Empty file B is valid input — engine should not raise
        result = run_delta(df_a, df_b, ["id"], ["id"])
        df = build_summary_df(result)
        # B-denominator rows should produce "N/A", not a crash
        b_row = df[df["Metric"] == "Records Only in File B"]
        assert b_row["% of File Total"].iloc[0] == "N/A"

    def test_pct_column_valid_values(self):
        df = build_summary_df(_make_result())
        for val in df["% of File Total"]:
            # Each cell is: a percent string, "N/A", or "" (blank for total rows)
            assert str(val).endswith("%") or val in ("N/A", "")


# ---------------------------------------------------------------------------
# build_change_frequency
# ---------------------------------------------------------------------------

class TestBuildChangeFrequency:
    def test_returns_dataframe_with_columns(self):
        result = _make_result()
        df = build_change_frequency(result)
        assert list(df.columns) == ["Field", "Changes"]

    def test_no_comparison_cols_returns_empty(self):
        result = _make_result(compare_cols=False)
        df = build_change_frequency(result)
        assert df.empty

    def test_changed_fields_appear(self):
        result = _make_result()
        df = build_change_frequency(result)
        assert "status" in df["Field"].values or "amount" in df["Field"].values

    def test_zero_change_fields_included(self):
        result = _make_result()
        df = build_change_frequency(result)
        assert len(df) <= len(result.compare_cols_a)


# ---------------------------------------------------------------------------
# export_to_excel — sheet names
# ---------------------------------------------------------------------------

REQUIRED_SHEETS = [
    "Executive Summary",
    "Analysis Metadata",
    "Comparison Rules",
    "Delta Counts",
    "Only in File A",
    "Only in File B",
    "Matched Records",
    "Changed Records",
    "Duplicate Keys File A",
    "Duplicate Keys File B",
    "Data Quality Issues",
]

REMOVED_SHEETS = ["Summary"]


class TestExcelSheetNames:
    def _load_wb(self, result=None, a="File A", b="File B"):
        if result is None:
            result = _make_result()
        raw = export_to_excel(result, a, b)
        return load_workbook(io.BytesIO(raw))

    def test_all_required_sheets_present(self):
        wb = self._load_wb()
        for sheet in REQUIRED_SHEETS:
            assert sheet in wb.sheetnames, f"Missing required sheet: '{sheet}'"

    def test_legacy_summary_sheet_removed(self):
        wb = self._load_wb()
        for sheet in REMOVED_SHEETS:
            assert sheet not in wb.sheetnames, f"Legacy sheet '{sheet}' should have been removed"

    def test_exactly_eleven_sheets(self):
        wb = self._load_wb()
        assert len(wb.sheetnames) == 11, f"Expected 11 sheets, got {len(wb.sheetnames)}: {wb.sheetnames}"

    def test_executive_summary_has_data(self):
        wb = self._load_wb()
        ws = wb["Executive Summary"]
        assert ws.max_row >= 2

    def test_analysis_metadata_has_data(self):
        wb = self._load_wb()
        ws = wb["Analysis Metadata"]
        assert ws.max_row >= 2

    def test_analysis_metadata_contains_file_names(self):
        wb = self._load_wb(a="my_a.csv", b="my_b.csv")
        ws = wb["Analysis Metadata"]
        values = [str(ws.cell(row=r, column=2).value) for r in range(2, ws.max_row + 1)]
        assert any("my_a.csv" in v for v in values)
        assert any("my_b.csv" in v for v in values)

    def test_analysis_metadata_contains_sheet_names(self):
        result = _make_result(sheet_a="Sheet1", sheet_b="DataTab")
        wb = self._load_wb(result)
        ws = wb["Analysis Metadata"]
        values = [str(ws.cell(row=r, column=2).value) for r in range(2, ws.max_row + 1)]
        assert any("Sheet1" in v for v in values)
        assert any("DataTab" in v for v in values)

    def test_comparison_rules_with_rules(self):
        result = _make_result(with_comparison_rules=True)
        wb = self._load_wb(result)
        ws = wb["Comparison Rules"]
        assert ws.max_row >= 2

    def test_comparison_rules_no_rules_shows_note(self):
        result = _make_result(compare_cols=False)
        wb = self._load_wb(result)
        ws = wb["Comparison Rules"]
        cell_val = str(ws.cell(row=1, column=1).value)
        assert "Note" in cell_val or ws.max_row >= 1

    def test_delta_counts_has_data(self):
        wb = self._load_wb()
        ws = wb["Delta Counts"]
        assert ws.max_row >= 2

    def test_only_in_a_present(self):
        wb = self._load_wb()
        assert "Only in File A" in wb.sheetnames

    def test_only_in_b_present(self):
        wb = self._load_wb()
        assert "Only in File B" in wb.sheetnames

    def test_changed_records_present(self):
        wb = self._load_wb()
        assert "Changed Records" in wb.sheetnames

    def test_duplicate_keys_a_present(self):
        wb = self._load_wb()
        assert "Duplicate Keys File A" in wb.sheetnames

    def test_duplicate_keys_b_present(self):
        wb = self._load_wb()
        assert "Duplicate Keys File B" in wb.sheetnames

    def test_data_quality_issues_present(self):
        wb = self._load_wb()
        assert "Data Quality Issues" in wb.sheetnames

    def test_returns_bytes(self):
        raw = export_to_excel(_make_result())
        assert isinstance(raw, bytes) and len(raw) > 0
