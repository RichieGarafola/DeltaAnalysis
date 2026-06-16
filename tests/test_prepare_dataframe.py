"""
Tests for src/io_utils.py: read_uploaded_file_raw / read_raw_file alias,
prepare_dataframe_from_raw, and preparation metadata in the Excel report.
"""
import io

import pandas as pd
import pytest
from openpyxl import load_workbook

from src.io_utils import prepare_dataframe_from_raw, read_raw_file, read_uploaded_file_raw
from src.delta_engine import run_delta
from src.reporting import export_to_excel


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

class _FakeFile:
    """Minimal stand-in for a Streamlit UploadedFile backed by bytes."""

    def __init__(self, content: bytes, name: str):
        self._buf = io.BytesIO(content)
        self.name = name

    def seek(self, pos: int) -> None:
        self._buf.seek(pos)

    def read(self, n: int = -1) -> bytes:
        return self._buf.read(n)

    def tell(self) -> int:
        return self._buf.tell()


def _csv_file(rows: list[list], name: str = "test.csv") -> _FakeFile:
    lines = [",".join(str(c) for c in row) for row in rows]
    content = "\n".join(lines).encode("utf-8")
    return _FakeFile(content, name)


def _make_raw_df(*rows) -> pd.DataFrame:
    """Build a raw (integer-column) DataFrame the same way read_uploaded_file_raw would."""
    data = [list(r) for r in rows]
    df = pd.DataFrame(data, dtype=str)
    df.columns = range(len(df.columns))
    return df


def _minimal_prep_meta(**overrides) -> dict:
    """Return a minimal valid prep metadata dict, optionally overriding any keys."""
    base = {
        "header_row_selected":        0,
        "drop_rows_above":            True,
        "drop_blank_rows":            True,
        "end_row_selected":           None,
        "rows_in_raw":                2,
        "rows_dropped_above_header":  0,
        "rows_dropped_blank":         0,
        "rows_dropped_after_end_row": 0,
        "rows_in_prepared":           2,
        "columns_in_prepared":        2,
        "blank_headers_renamed":      0,
        "duplicate_headers_renamed":  0,
    }
    base.update(overrides)
    return base


# ---------------------------------------------------------------------------
# TestReadRawFile - alias and primary name both work
# ---------------------------------------------------------------------------

class TestReadRawFile:
    def test_csv_returns_integer_columns(self):
        f = _csv_file([["id", "name"], ["1", "Alice"], ["2", "Bob"]])
        df = read_uploaded_file_raw(f)
        assert list(df.columns) == [0, 1]

    def test_alias_read_raw_file_works(self):
        f = _csv_file([["id"], ["1"]])
        df = read_raw_file(f)
        assert list(df.columns) == [0]

    def test_csv_first_row_is_data_not_header(self):
        f = _csv_file([["id", "name"], ["1", "Alice"]])
        df = read_uploaded_file_raw(f)
        assert df.iloc[0, 0] == "id"

    def test_csv_all_values_are_strings(self):
        f = _csv_file([["col"], ["42"], ["3.14"]])
        df = read_uploaded_file_raw(f)
        assert all(isinstance(v, str) for v in df.iloc[:, 0])

    def test_csv_blank_cells_become_empty_string(self):
        content = b"a,,c\n1,,3\n"
        f = _FakeFile(content, "test.csv")
        df = read_uploaded_file_raw(f)
        assert df.iloc[1, 1] == ""

    def test_seek_restored_after_read(self):
        f = _csv_file([["x"], ["1"]])
        read_uploaded_file_raw(f)
        assert f._buf.tell() == 0

    def test_none_file_raises_value_error(self):
        with pytest.raises(ValueError, match="No file"):
            read_uploaded_file_raw(None)

    def test_unsupported_extension_raises_value_error(self):
        f = _FakeFile(b"data", "data.txt")
        with pytest.raises(ValueError, match="Unsupported file format"):
            read_uploaded_file_raw(f)

    def test_row_count_matches_content(self):
        f = _csv_file([["h1", "h2"], ["a", "b"], ["c", "d"], ["e", "f"]])
        df = read_uploaded_file_raw(f)
        assert len(df) == 4

    def test_column_count_matches_content(self):
        f = _csv_file([["a", "b", "c", "d"]])
        df = read_uploaded_file_raw(f)
        assert len(df.columns) == 4


# ---------------------------------------------------------------------------
# TestPrepareDataframeFromRaw
# ---------------------------------------------------------------------------

class TestPrepareDataframeFromRaw:
    def _simple_raw(self):
        return _make_raw_df(
            ["id", "name", "amount"],
            ["1",  "Alice", "100"],
            ["2",  "Bob",   "200"],
            ["3",  "Carol", "300"],
        )

    def test_header_row_0_default(self):
        df, meta, warns = prepare_dataframe_from_raw(self._simple_raw())
        assert list(df.columns) == ["id", "name", "amount"]
        assert len(df) == 3
        assert warns == []

    def test_header_row_index_2(self):
        raw = _make_raw_df(
            ["Report Title", "", ""],
            ["Agency: DoD",  "", ""],
            ["id", "name",   "amount"],
            ["1",  "Alice",  "100"],
        )
        df, meta, _ = prepare_dataframe_from_raw(raw, header_row_index=2)
        assert list(df.columns) == ["id", "name", "amount"]
        assert len(df) == 1
        assert df.iloc[0]["id"] == "1"

    def test_rows_above_dropped_by_default(self):
        raw = _make_raw_df(
            ["metadata", ""],
            ["id", "name"],
            ["1", "Alice"],
        )
        df, meta, _ = prepare_dataframe_from_raw(raw, header_row_index=1, drop_rows_above=True)
        assert len(df) == 1
        assert meta["rows_dropped_above_header"] == 1

    def test_rows_above_kept_when_disabled(self):
        raw = _make_raw_df(
            ["metadata", "notes"],
            ["id", "name"],
            ["1", "Alice"],
        )
        df, meta, _ = prepare_dataframe_from_raw(raw, header_row_index=1, drop_rows_above=False)
        assert len(df) == 2
        assert meta["rows_dropped_above_header"] == 0

    def test_blank_rows_dropped_by_default(self):
        raw = _make_raw_df(
            ["id", "name"],
            ["1", "Alice"],
            ["",  ""],
            ["2", "Bob"],
        )
        df, meta, _ = prepare_dataframe_from_raw(raw)
        assert len(df) == 2
        assert meta["rows_dropped_blank"] == 1

    def test_blank_rows_kept_when_disabled(self):
        raw = _make_raw_df(
            ["id", "name"],
            ["1", "Alice"],
            ["",  ""],
        )
        df, meta, _ = prepare_dataframe_from_raw(raw, drop_blank_rows=False)
        assert len(df) == 2
        assert meta["rows_dropped_blank"] == 0

    def test_end_row_filter(self):
        raw = _make_raw_df(
            ["id"],
            ["1"],
            ["2"],
            ["3"],
            ["Total", "10"],
        )
        # end_row_index=3 (0-based inclusive) → includes data rows 1, 2, 3
        df, meta, _ = prepare_dataframe_from_raw(raw, end_row_index=3)
        assert len(df) == 3
        assert meta["end_row_selected"] == 3
        assert meta["rows_dropped_after_end_row"] == 1

    def test_end_row_none_includes_all(self):
        raw = self._simple_raw()
        df, meta, _ = prepare_dataframe_from_raw(raw, end_row_index=None)
        assert len(df) == 3
        assert meta["end_row_selected"] is None
        assert meta["rows_dropped_after_end_row"] == 0

    def test_end_row_before_header_ignored_with_warning(self):
        raw = _make_raw_df(["id"], ["1"], ["2"])
        df, meta, warns = prepare_dataframe_from_raw(
            raw, header_row_index=1, end_row_index=0
        )
        assert len(df) == 1
        assert any("ignored" in w.lower() for w in warns)

    def test_end_row_out_of_bounds_ignored_with_warning(self):
        raw = _make_raw_df(["id"], ["1"])
        df, meta, warns = prepare_dataframe_from_raw(raw, end_row_index=99)
        assert len(df) == 1
        assert any("exceeds" in w.lower() for w in warns)

    def test_out_of_bounds_header_raises(self):
        raw = _make_raw_df(["id"], ["1"])
        with pytest.raises(ValueError, match="out of range"):
            prepare_dataframe_from_raw(raw, header_row_index=99)

    def test_empty_dataframe_raises(self):
        empty = pd.DataFrame()
        with pytest.raises(ValueError, match="empty"):
            prepare_dataframe_from_raw(empty)

    def test_blank_header_renamed_to_unnamed(self):
        raw = _make_raw_df(
            ["id", "", "amount"],
            ["1",  "X", "100"],
        )
        df, meta, warns = prepare_dataframe_from_raw(raw)
        assert "Unnamed_1" in df.columns
        assert any("blank header" in w.lower() for w in warns)
        assert meta["blank_headers_renamed"] == 1

    def test_duplicate_headers_renamed(self):
        raw = _make_raw_df(
            ["id", "val", "val"],
            ["1",  "a",   "b"],
        )
        df, meta, warns = prepare_dataframe_from_raw(raw)
        assert "val" in df.columns
        assert "val_2" in df.columns
        assert any("Duplicate" in w for w in warns)
        assert meta["duplicate_headers_renamed"] == 1

    def test_whitespace_stripped_from_headers(self):
        raw = _make_raw_df(
            ["  id  ", " name ", "amount"],
            ["1", "Alice", "100"],
        )
        df, _, _ = prepare_dataframe_from_raw(raw)
        assert "id" in df.columns
        assert "name" in df.columns

    def test_metadata_structure(self):
        _, meta, _ = prepare_dataframe_from_raw(self._simple_raw())
        required_keys = {
            "header_row_selected",
            "drop_rows_above",
            "drop_blank_rows",
            "end_row_selected",
            "rows_in_raw",
            "rows_dropped_above_header",
            "rows_dropped_blank",
            "rows_dropped_after_end_row",
            "rows_in_prepared",
            "columns_in_prepared",
            "blank_headers_renamed",
            "duplicate_headers_renamed",
        }
        assert required_keys == set(meta.keys())

    def test_metadata_values_correct(self):
        raw = _make_raw_df(
            ["ignore", ""],
            ["id", "name"],
            ["1", "Alice"],
            ["",  ""],
        )
        df, meta, _ = prepare_dataframe_from_raw(raw, header_row_index=1)
        assert meta["header_row_selected"] == 1
        assert meta["rows_dropped_above_header"] == 1
        assert meta["rows_dropped_blank"] == 1
        assert meta["rows_in_prepared"] == 1
        assert meta["columns_in_prepared"] == 2

    def test_returns_tuple_of_three(self):
        result = prepare_dataframe_from_raw(self._simple_raw())
        assert len(result) == 3
        df, meta, warns = result
        assert isinstance(df, pd.DataFrame)
        assert isinstance(meta, dict)
        assert isinstance(warns, list)

    def test_empty_result_adds_warning(self):
        raw = _make_raw_df(["id"], ["", ""])
        df, meta, warns = prepare_dataframe_from_raw(raw, drop_blank_rows=True)
        assert df.empty
        assert any("no rows" in w.lower() for w in warns)


# ---------------------------------------------------------------------------
# TestPreparationMetadataInReport
# ---------------------------------------------------------------------------

class TestPreparationMetadataInReport:
    def _make_result(self):
        df_a = pd.DataFrame({"id": ["1", "2"], "val": ["a", "b"]})
        df_b = pd.DataFrame({"id": ["1", "2"], "val": ["a", "c"]})
        return run_delta(df_a, df_b, ["id"], ["id"], ["val"], ["val"])

    def _load_metadata_values(self, wb) -> list[str]:
        ws = wb["Analysis Metadata"]
        return [str(ws.cell(row=r, column=2).value) for r in range(2, ws.max_row + 1)]

    def _load_metadata_params(self, wb) -> list[str]:
        ws = wb["Analysis Metadata"]
        return [str(ws.cell(row=r, column=1).value) for r in range(2, ws.max_row + 1)]

    def test_metadata_without_prep_meta_has_no_prep_rows(self):
        result = self._make_result()
        raw = export_to_excel(result)
        wb = load_workbook(io.BytesIO(raw))
        params = self._load_metadata_params(wb)
        assert not any("Header Row Selected" in p for p in params)

    def test_metadata_with_source_prep_meta_includes_source_rows(self):
        result = self._make_result()
        prep_a = _minimal_prep_meta(
            header_row_selected=2,
            rows_dropped_above_header=2,
            rows_dropped_blank=1,
            rows_in_prepared=5,
            columns_in_prepared=3,
        )
        raw = export_to_excel(result, source_prep_metadata=prep_a)
        wb = load_workbook(io.BytesIO(raw))
        params = self._load_metadata_params(wb)
        assert any("Source Header Row Selected" in p for p in params)
        assert any("Source Rows Dropped Above Header" in p for p in params)
        assert any("Source Blank Rows Dropped" in p for p in params)

    def test_metadata_with_comparison_prep_meta_includes_comparison_rows(self):
        result = self._make_result()
        prep_b = _minimal_prep_meta(
            rows_dropped_blank=3,
            rows_in_prepared=10,
            columns_in_prepared=4,
        )
        raw = export_to_excel(result, comparison_prep_metadata=prep_b)
        wb = load_workbook(io.BytesIO(raw))
        params = self._load_metadata_params(wb)
        assert any("Comparison Header Row Selected" in p for p in params)
        assert any("Comparison Blank Rows Dropped" in p for p in params)

    def test_end_row_selected_always_present_in_prep_rows(self):
        result = self._make_result()
        prep_no_end = _minimal_prep_meta(end_row_selected=None)
        prep_with_end = _minimal_prep_meta(end_row_selected=50)

        wb_no = load_workbook(io.BytesIO(export_to_excel(result, source_prep_metadata=prep_no_end)))
        wb_yes = load_workbook(io.BytesIO(export_to_excel(result, source_prep_metadata=prep_with_end)))

        params_no = self._load_metadata_params(wb_no)
        values_no = self._load_metadata_values(wb_no)
        params_yes = self._load_metadata_params(wb_yes)
        values_yes = self._load_metadata_values(wb_yes)

        # Row always present
        assert any("Source End Row Selected" in p for p in params_no)
        assert any("Source End Row Selected" in p for p in params_yes)

        # Value reflects whether it was set
        idx_no = next(i for i, p in enumerate(params_no) if "Source End Row Selected" in p)
        idx_yes = next(i for i, p in enumerate(params_yes) if "Source End Row Selected" in p)
        assert values_no[idx_no] == "Not applied"
        assert values_yes[idx_yes] == "51"  # 50 (0-based) + 1 = 51 displayed

    def test_header_row_displayed_as_one_based(self):
        result = self._make_result()
        prep_a = _minimal_prep_meta(header_row_selected=2)  # 0-based 2 → 1-based 3
        raw = export_to_excel(result, source_prep_metadata=prep_a)
        wb = load_workbook(io.BytesIO(raw))
        params = self._load_metadata_params(wb)
        values = self._load_metadata_values(wb)
        idx = next(i for i, p in enumerate(params) if "Source Header Row Selected" in p)
        assert values[idx] == "3"

    def test_prep_meta_counts_written_to_excel(self):
        result = self._make_result()
        prep_a = _minimal_prep_meta(
            header_row_selected=2,   # displayed as "3"
            rows_dropped_blank=7,
        )
        raw = export_to_excel(result, source_prep_metadata=prep_a)
        wb = load_workbook(io.BytesIO(raw))
        values = self._load_metadata_values(wb)
        assert "3" in values   # 1-based header row
        assert "7" in values   # blank rows dropped
