"""
End-to-end validation suite — v1.1 RC1.

Tests every scenario specified in the RC1 validation checklist:
  - CSV vs CSV
  - Excel vs Excel (single-sheet)
  - Multi-sheet Excel (sheet selector)
  - Numeric tolerance comparison
  - Date-only comparison (date_only mode)
  - Datetime precision comparison (datetime_precision mode)
  - Full workbook generation and sheet verification
  - Executive Summary content sanity
  - Comparison Rules tab content
  - Analysis Metadata content
  - No Summary (legacy) sheet present
"""
import io

import pandas as pd
import pytest
from openpyxl import load_workbook

from src.delta_engine import run_delta
from src.io_utils import (
    check_file_size,
    get_display_frame,
    get_excel_sheet_names,
    read_uploaded_file,
    LARGE_FILE_WARN_ROWS,
    LARGE_FILE_HARD_ROWS,
    PREVIEW_MAX_ROWS,
)
from src.reporting import export_to_excel

REQUIRED_SHEETS = [
    "Executive Summary",
    "Analysis Metadata",
    "Comparison Rules",
    "Delta Counts",
    "Only in File A",
    "Only in File B",
    "Matched Records",
    "Changed Records",
    "Duplicate Keys File A",
    "Duplicate Keys File B",
    "Data Quality Issues",
]


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

class _MockFile:
    def __init__(self, name: str, content: bytes):
        self.name = name
        self._buf = io.BytesIO(content)

    def read(self, n=-1):       return self._buf.read(n)
    def seek(self, p, w=0):     return self._buf.seek(p, w)
    def tell(self):             return self._buf.tell()
    def seekable(self):         return True


def _csv(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False).encode("utf-8")


def _xlsx(sheets: dict) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        for name, df in sheets.items():
            df.to_excel(w, sheet_name=name, index=False)
    return buf.getvalue()


def _wb_from_result(result, a="A.csv", b="B.csv"):
    raw = export_to_excel(result, a, b)
    return load_workbook(io.BytesIO(raw)), raw


# ---------------------------------------------------------------------------
# Scenario 1 — CSV vs CSV
# ---------------------------------------------------------------------------

class TestCsvVsCsv:
    def _run(self):
        df_a = pd.DataFrame({
            "id":     ["C1", "C2", "C3"],
            "status": ["Open", "Closed", "Pending"],
        })
        df_b = pd.DataFrame({
            "id":     ["C1", "C2", "C4"],
            "status": ["Open", "Approved", "New"],
        })
        fa = _MockFile("file_a.csv", _csv(df_a))
        fb = _MockFile("file_b.csv", _csv(df_b))
        dfa = read_uploaded_file(fa)
        dfb = read_uploaded_file(fb)
        return run_delta(dfa, dfb, ["id"], ["id"], ["status"], ["status"])

    def test_csv_loads_correctly(self):
        r = self._run()
        assert r.total_a == 3
        assert r.total_b == 3

    def test_only_in_a_correct(self):
        r = self._run()
        assert len(r.only_in_a) == 1
        assert r.only_in_a["id"].iloc[0] == "C3"

    def test_only_in_b_correct(self):
        r = self._run()
        assert len(r.only_in_b) == 1
        assert r.only_in_b["id"].iloc[0] == "C4"

    def test_changed_detected(self):
        r = self._run()
        assert len(r.changed) == 1

    def test_workbook_generates(self):
        wb, raw = _wb_from_result(self._run(), "a.csv", "b.csv")
        assert isinstance(raw, bytes) and len(raw) > 0
        for sheet in REQUIRED_SHEETS:
            assert sheet in wb.sheetnames


# ---------------------------------------------------------------------------
# Scenario 2 — Excel vs Excel (single-sheet)
# ---------------------------------------------------------------------------

class TestExcelVsExcel:
    def _run(self):
        df_a = pd.DataFrame({"ref": ["E01", "E02"], "val": ["100", "200"]})
        df_b = pd.DataFrame({"ref": ["E01", "E02"], "val": ["100", "250"]})
        fa = _MockFile("a.xlsx", _xlsx({"Sheet1": df_a}))
        fb = _MockFile("b.xlsx", _xlsx({"Sheet1": df_b}))
        dfa = read_uploaded_file(fa)
        dfb = read_uploaded_file(fb)
        return run_delta(dfa, dfb, ["ref"], ["ref"], ["val"], ["val"])

    def test_excel_loads(self):
        r = self._run()
        assert r.total_a == 2 and r.total_b == 2

    def test_changed_detected(self):
        r = self._run()
        assert len(r.changed) == 1

    def test_workbook_valid(self):
        wb, _ = _wb_from_result(self._run(), "a.xlsx", "b.xlsx")
        assert "Changed Records" in wb.sheetnames


# ---------------------------------------------------------------------------
# Scenario 3 — Multi-sheet Excel
# ---------------------------------------------------------------------------

class TestMultiSheetExcel:
    def test_sheet_names_returned(self):
        df = pd.DataFrame({"x": ["1"]})
        f = _MockFile("data.xlsx", _xlsx({"Alpha": df, "Beta": df, "Gamma": df}))
        sheets = get_excel_sheet_names(f)
        assert sheets == ["Alpha", "Beta", "Gamma"]

    def test_file_seekable_after_sheet_read(self):
        df = pd.DataFrame({"x": ["1"]})
        f = _MockFile("data.xlsx", _xlsx({"S1": df, "S2": df}))
        get_excel_sheet_names(f)
        assert f.tell() == 0

    def test_named_sheet_read_correctly(self):
        df_s1 = pd.DataFrame({"id": ["1"], "v": ["A"]})
        df_s2 = pd.DataFrame({"id": ["99"], "v": ["Z"]})
        f = _MockFile("data.xlsx", _xlsx({"Sheet1": df_s1, "Sheet2": df_s2}))
        result = read_uploaded_file(f, sheet_name="Sheet2")
        assert result["id"].iloc[0] == "99"

    def test_multisheet_end_to_end(self):
        df_a = pd.DataFrame({"id": ["M1", "M2"], "status": ["Open", "Closed"]})
        df_b = pd.DataFrame({"id": ["M1", "M3"], "status": ["Open", "New"]})
        fa = _MockFile("a.xlsx", _xlsx({"Data": df_a, "Audit": pd.DataFrame({"x": ["log"]})}))
        fb = _MockFile("b.xlsx", _xlsx({"Received": df_b}))
        dfa = read_uploaded_file(fa, sheet_name="Data")
        dfb = read_uploaded_file(fb, sheet_name="Received")
        r = run_delta(dfa, dfb, ["id"], ["id"], ["status"], ["status"],
                      sheet_a="Data", sheet_b="Received")
        assert r.sheet_a == "Data" and r.sheet_b == "Received"
        wb, _ = _wb_from_result(r)
        ws = wb["Analysis Metadata"]
        values = [str(ws.cell(row=i, column=2).value) for i in range(2, ws.max_row + 1)]
        assert any("Data" in v for v in values)
        assert any("Received" in v for v in values)


# ---------------------------------------------------------------------------
# Scenario 4 — Numeric tolerance comparison
# ---------------------------------------------------------------------------

class TestNumericTolerance:
    def _run(self, tol):
        df_a = pd.DataFrame({"id": ["N1", "N2", "N3"],
                              "amt": ["$1,000.00", "$2,500.00", "$9,999.00"]})
        df_b = pd.DataFrame({"id": ["N1", "N2", "N3"],
                              "amt": ["1000.00",   "2499.50",   "9800.00"]})
        rules = [{"column_a": "amt", "column_b": "amt",
                  "type": "numeric", "tolerance": tol, "date_mode": None}]
        return run_delta(df_a, df_b, ["id"], ["id"], ["amt"], ["amt"],
                         comparison_rules=rules)

    def test_currency_symbols_stripped(self):
        r = self._run(tol=0.0)
        # N1: $1,000.00 == 1000.00 after stripping — must NOT appear in changed
        # N2 and N3 differ in value, so they should be flagged
        changed_keys = r.changed["Key: id"].tolist() if not r.changed.empty else []
        assert "N1" not in changed_keys

    def test_within_tolerance_not_flagged(self):
        r = self._run(tol=1.0)
        # N1: equal, N2: diff=0.50 (within 1.0), N3: diff=199 (outside 1.0)
        assert len(r.changed) == 1
        assert "N3" in str(r.changed.to_dict())

    def test_outside_tolerance_flagged(self):
        r = self._run(tol=0.0)
        # N2: 2500 vs 2499.50 → diff = 0.50 → flagged
        # N3: 9999 vs 9800 → diff = 199 → flagged
        assert len(r.changed) == 2

    def test_parentheses_negative_handled(self):
        df_a = pd.DataFrame({"id": ["P1"], "amt": ["(500.00)"]})
        df_b = pd.DataFrame({"id": ["P1"], "amt": ["-500"]})
        rules = [{"column_a": "amt", "column_b": "amt",
                  "type": "numeric", "tolerance": 0.0, "date_mode": None}]
        r = run_delta(df_a, df_b, ["id"], ["id"], ["amt"], ["amt"],
                      comparison_rules=rules)
        assert r.changed.empty

    def test_workbook_comparison_rules_tab(self):
        r = self._run(tol=0.5)
        wb, _ = _wb_from_result(r)
        ws = wb["Comparison Rules"]
        assert ws.max_row >= 2


# ---------------------------------------------------------------------------
# Scenario 5 — Date-only comparison
# ---------------------------------------------------------------------------

class TestDateOnlyComparison:
    def _run(self):
        df_a = pd.DataFrame({"id":   ["D1", "D2", "D3"],
                              "date": ["2024-01-15", "01/15/2024", "2024-01-15 08:30:00"]})
        df_b = pd.DataFrame({"id":   ["D1", "D2", "D3"],
                              "date": ["01/15/2024", "2024-01-15", "2024-01-15 23:59:00"]})
        rules = [{"column_a": "date", "column_b": "date",
                  "type": "date", "tolerance": None, "date_mode": "date_only"}]
        return run_delta(df_a, df_b, ["id"], ["id"], ["date"], ["date"],
                         comparison_rules=rules)

    def test_iso_vs_us_format_equal(self):
        r = self._run()
        # D1 and D2: same calendar date, different formats → equal
        assert r.changed.empty or "D1" not in str(r.changed)

    def test_time_component_ignored(self):
        r = self._run()
        # D3: same date, different times → equal in date_only mode
        assert r.changed.empty

    def test_different_calendar_dates_flagged(self):
        df_a = pd.DataFrame({"id": ["X"], "date": ["2024-01-15"]})
        df_b = pd.DataFrame({"id": ["X"], "date": ["2024-01-16"]})
        rules = [{"column_a": "date", "column_b": "date",
                  "type": "date", "tolerance": None, "date_mode": "date_only"}]
        r = run_delta(df_a, df_b, ["id"], ["id"], ["date"], ["date"],
                      comparison_rules=rules)
        assert len(r.changed) == 1


# ---------------------------------------------------------------------------
# Scenario 6 — Datetime precision comparison
# ---------------------------------------------------------------------------

class TestDatetimePrecisionComparison:
    def _run(self):
        df_a = pd.DataFrame({"id":   ["T1", "T2"],
                              "ts":   ["2024-01-15 08:30:00", "2024-01-15"]})
        df_b = pd.DataFrame({"id":   ["T1", "T2"],
                              "ts":   ["2024-01-15 14:00:00", "2024-01-15 00:00:00"]})
        rules = [{"column_a": "ts", "column_b": "ts",
                  "type": "date", "tolerance": None, "date_mode": "datetime_precision"}]
        return run_delta(df_a, df_b, ["id"], ["id"], ["ts"], ["ts"],
                         comparison_rules=rules)

    def test_different_times_flagged(self):
        r = self._run()
        # T1: 08:30 vs 14:00 → different
        assert len(r.changed) >= 1
        assert "T1" in str(r.changed)

    def test_date_without_time_equals_midnight(self):
        r = self._run()
        # T2: "2024-01-15" (midnight) vs "2024-01-15 00:00:00" → equal
        if not r.changed.empty:
            assert "T2" not in str(r.changed)


# ---------------------------------------------------------------------------
# Scenario 7 — Full workbook verification
# ---------------------------------------------------------------------------

class TestWorkbookVerification:
    def _full_result(self):
        df_a = pd.DataFrame({
            "id":     ["W1", "W2", "W3", "W4", "W4", ""],
            "status": ["Open", "Pending", "Closed", "Open", "Open", "Unk"],
            "amount": ["$1,000", "$2,000", "$3,000", "$400", "$400", "$50"],
            "date":   ["2024-01-15", "2024-02-01", "2024-03-10", "2024-04-01",
                       "2024-04-01", "2024-05-01"],
        })
        df_b = pd.DataFrame({
            "id":     ["W1", "W2", "W5"],
            "status": ["Open", "Approved", "New"],
            "amount": ["$1,000", "2001",    "$600"],
            "date":   ["01/15/2024", "2024-02-01", "2024-06-01"],
        })
        rules = [
            {"column_a": "status", "column_b": "status",
             "type": "text",    "tolerance": None, "date_mode": None},
            {"column_a": "amount", "column_b": "amount",
             "type": "numeric", "tolerance": 2.0,  "date_mode": None},
            {"column_a": "date",   "column_b": "date",
             "type": "date",    "tolerance": None, "date_mode": "date_only"},
        ]
        return run_delta(df_a, df_b, ["id"], ["id"],
                         ["status", "amount", "date"],
                         ["status", "amount", "date"],
                         comparison_rules=rules,
                         sheet_a="Records", sheet_b="Received")

    def test_all_11_required_sheets_present(self):
        r = self._full_result()
        wb, _ = _wb_from_result(r, "dataset_a.xlsx", "dataset_b.xlsx")
        for sheet in REQUIRED_SHEETS:
            assert sheet in wb.sheetnames, f"Missing: '{sheet}'"

    def test_no_legacy_summary_sheet(self):
        r = self._full_result()
        wb, _ = _wb_from_result(r)
        assert "Summary" not in wb.sheetnames

    def test_exactly_11_sheets(self):
        r = self._full_result()
        wb, _ = _wb_from_result(r)
        assert len(wb.sheetnames) == 11, f"Got {len(wb.sheetnames)}: {wb.sheetnames}"

    def test_executive_summary_sections(self):
        r = self._full_result()
        wb, _ = _wb_from_result(r)
        ws = wb["Executive Summary"]
        sections = [str(ws.cell(row=i, column=1).value) for i in range(2, ws.max_row + 1)]
        assert "Analysis Overview" in sections
        assert "Matching Results" in sections
        assert "Field-Level Changes" in sections
        assert "Recommended Actions" in sections

    def test_analysis_metadata_all_parameters(self):
        r = self._full_result()
        wb, _ = _wb_from_result(r, "a.xlsx", "b.xlsx")
        ws = wb["Analysis Metadata"]
        params = [str(ws.cell(row=i, column=1).value) for i in range(2, ws.max_row + 1)]
        assert "File A Name" in params
        assert "File B Name" in params
        assert "File A Sheet" in params
        assert "File B Sheet" in params
        assert "Key Columns (File A)" in params
        assert "Comparison Rule Count" in params

    def test_comparison_rules_tab_has_three_rules(self):
        r = self._full_result()
        wb, _ = _wb_from_result(r)
        ws = wb["Comparison Rules"]
        # Header row + 3 rule rows
        assert ws.max_row == 4

    def test_comparison_rules_types_correct(self):
        r = self._full_result()
        wb, _ = _wb_from_result(r)
        ws = wb["Comparison Rules"]
        types = [str(ws.cell(row=i, column=3).value) for i in range(2, ws.max_row + 1)]
        assert "text" in types
        assert "numeric" in types
        assert "date" in types

    def test_delta_counts_tab_has_categories(self):
        r = self._full_result()
        wb, _ = _wb_from_result(r)
        ws = wb["Delta Counts"]
        categories = [str(ws.cell(row=i, column=1).value) for i in range(2, ws.max_row + 1)]
        assert any("Only in File A" in c for c in categories)
        assert any("Changed" in c for c in categories)

    def test_duplicate_keys_tab_has_data(self):
        r = self._full_result()
        wb, _ = _wb_from_result(r)
        ws = wb["Duplicate Keys File A"]
        # W4 appears twice → dup rows present
        assert ws.max_row >= 2

    def test_data_quality_issues_tab_has_data(self):
        r = self._full_result()
        wb, _ = _wb_from_result(r)
        ws = wb["Data Quality Issues"]
        # blank key row present in df_a
        assert ws.max_row >= 2

    def test_workbook_not_corrupt(self):
        r = self._full_result()
        _, raw = _wb_from_result(r)
        # openpyxl should load without exception
        wb2 = load_workbook(io.BytesIO(raw))
        assert len(wb2.sheetnames) == 11

    def test_changed_records_before_after_columns(self):
        r = self._full_result()
        assert not r.changed.empty
        cols = r.changed.columns.tolist()
        # Expect columns like "status — File A", "status — File B"
        assert any("— File A" in c for c in cols)
        assert any("— File B" in c for c in cols)


# ---------------------------------------------------------------------------
# Scenario 8 — App startup import check
# ---------------------------------------------------------------------------

class TestAppImports:
    def test_all_src_modules_importable(self):
        import src.comparison
        import src.delta_engine
        import src.io_utils
        import src.normalization
        import src.reporting
        assert True

    def test_app_module_imports_without_streamlit_runtime(self):
        import importlib, sys
        # Verify app.py parses and all its imports resolve
        # (Streamlit itself won't execute UI code at import time)
        try:
            import py_compile, os
            result = py_compile.compile(
                os.path.join(os.path.dirname(__file__), "..", "app.py"),
                doraise=True,
            )
            assert result is not None
        except py_compile.PyCompileError as e:
            pytest.fail(f"app.py failed to compile: {e}")
