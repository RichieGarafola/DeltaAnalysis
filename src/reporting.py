"""
Excel reporting module.

Builds an audit-grade, multi-tab Excel workbook from a DeltaResult.
Includes an Executive Narrative tab with auto-generated plain-English
summary language suitable for leadership briefings.
"""
from __future__ import annotations

import io
from datetime import datetime
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from src.delta_engine import DeltaResult


# ---------------------------------------------------------------------------
# Color palette
# ---------------------------------------------------------------------------

COLORS = {
    "navy":        "1F4E79",
    "med_blue":    "2F75B5",
    "light_blue":  "9DC3E6",
    "pale_blue":   "D9EAF7",
    "slate":       "44546A",
    "light_gray":  "F2F2F2",
    "white":       "FFFFFF",
}

_HEADER_COLORS = {
    "Executive Summary":                COLORS["slate"],
    "Analysis Metadata":                COLORS["slate"],
    "Comparison Rules":                 COLORS["slate"],
    "Delta Counts":                     COLORS["navy"],
    "Source Only Records":              COLORS["navy"],
    "Comparison Only Records":          COLORS["med_blue"],
    "Matched Records":                  COLORS["navy"],
    "Changed Records":                  COLORS["med_blue"],
    "Source Duplicate Keys":            COLORS["slate"],
    "Comparison Duplicate Keys":        COLORS["slate"],
    "Data Quality Flags":               COLORS["slate"],
}

_ROW_FILL_COLORS = {
    "Source Only Records":              COLORS["pale_blue"],
    "Comparison Only Records":          COLORS["pale_blue"],
    "Changed Records":                  COLORS["pale_blue"],
    "Source Duplicate Keys":            COLORS["light_gray"],
    "Comparison Duplicate Keys":        COLORS["light_gray"],
}


# ---------------------------------------------------------------------------
# Public: summary DataFrame
# ---------------------------------------------------------------------------

def build_summary_df(result: DeltaResult) -> pd.DataFrame:
    """Produce a human-readable summary statistics table."""
    total_a = result.total_a
    total_b = result.total_b
    n_matched = len(result.matched)

    def pct(n: int, denom: int) -> str:
        if denom == 0:
            return "N/A"
        return f"{n / denom * 100:.1f}%"

    rows = [
        ("Source Dataset: Total Records",          total_a,                   ""),
        ("Comparison Dataset: Total Records",      total_b,                   ""),
        ("Source Only Records",                    len(result.only_in_a),     pct(len(result.only_in_a), total_a)),
        ("Comparison Only Records",                len(result.only_in_b),     pct(len(result.only_in_b), total_b)),
        ("Matched Records",                        n_matched,                 pct(n_matched, total_a)),
        ("Changed Records",                        len(result.changed),       pct(len(result.changed), n_matched) if n_matched else "N/A"),
        ("Source Duplicate Keys",                  len(result.duplicates_a),  pct(len(result.duplicates_a), total_a)),
        ("Comparison Duplicate Keys",              len(result.duplicates_b),  pct(len(result.duplicates_b), total_b)),
        ("Source Missing Keys",                    len(result.blank_keys_a),  pct(len(result.blank_keys_a), total_a)),
        ("Comparison Missing Keys",                len(result.blank_keys_b),  pct(len(result.blank_keys_b), total_b)),
    ]

    return pd.DataFrame(rows, columns=["Metric", "Count", "% of Dataset Total"])


# ---------------------------------------------------------------------------
# Public: field-change frequency
# ---------------------------------------------------------------------------

def build_change_frequency(result: DeltaResult) -> pd.DataFrame:
    """Count how many times each compared field changed."""
    if result.changed.empty or not result.compare_cols_a:
        return pd.DataFrame(columns=["Field", "Changes"])

    rows = []
    for col in result.compare_cols_a:
        a_col = f"{col} - Source"
        b_col = f"{col} - Comparison"
        if a_col in result.changed.columns and b_col in result.changed.columns:
            n_changed = (result.changed[a_col] != result.changed[b_col]).sum()
            rows.append({"Field": col, "Changes": int(n_changed)})

    if not rows:
        return pd.DataFrame(columns=["Field", "Changes"])
    df = pd.DataFrame(rows).sort_values("Changes", ascending=False).reset_index(drop=True)
    return df


# ---------------------------------------------------------------------------
# Public: Excel export
# ---------------------------------------------------------------------------

def export_to_excel(
    result: DeltaResult,
    file_a_name: str = "Source Dataset",
    file_b_name: str = "Comparison Dataset",
    prep_meta_a: Optional[dict] = None,
    prep_meta_b: Optional[dict] = None,
) -> bytes:
    """
    Build a polished, multi-tab Excel workbook from a DeltaResult.
    Returns raw bytes for Streamlit download.

    prep_meta_a / prep_meta_b : optional dicts from prepare_dataframe_from_raw;
    when provided, their contents are appended to the Analysis Metadata tab.
    """
    run_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")

    narrative    = _build_narrative(result, file_a_name, file_b_name, run_timestamp)
    delta_counts = _build_delta_counts_df(result)

    # Data Quality Flags: rows excluded due to missing match key identifiers
    dq_rows: list[dict] = []
    for _, row in result.blank_keys_a.iterrows():
        dq_rows.append({"Dataset": "Source", "Filename": file_a_name, "Flag": "Missing Identifier", **row.to_dict()})
    for _, row in result.blank_keys_b.iterrows():
        dq_rows.append({"Dataset": "Comparison", "Filename": file_b_name, "Flag": "Missing Identifier", **row.to_dict()})
    dq_df = pd.DataFrame(dq_rows) if dq_rows else pd.DataFrame(columns=["Dataset", "Filename", "Flag"])

    metadata_df = _build_metadata_df(result, file_a_name, file_b_name, run_timestamp, prep_meta_a, prep_meta_b)
    rules_df    = _build_rules_df(result)

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        _write_narrative(narrative, writer)
        metadata_df.to_excel(writer,            sheet_name="Analysis Metadata",     index=False)
        rules_df.to_excel(writer,               sheet_name="Comparison Rules",      index=False)
        delta_counts.to_excel(writer,           sheet_name="Delta Counts",          index=False)
        _write_sheet(result.only_in_a,    writer, "Source Only Records")
        _write_sheet(result.only_in_b,    writer, "Comparison Only Records")
        _write_sheet(result.matched,      writer, "Matched Records")
        _write_sheet(result.changed,      writer, "Changed Records")
        _write_sheet(result.duplicates_a, writer, "Source Duplicate Keys")
        _write_sheet(result.duplicates_b, writer, "Comparison Duplicate Keys")
        _write_sheet(dq_df,               writer, "Data Quality Flags")

    # Post-process: styling
    buffer.seek(0)
    wb = load_workbook(buffer)
    _apply_styles(wb)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ---------------------------------------------------------------------------
# Narrative builder
# ---------------------------------------------------------------------------

def _build_narrative(
    result: DeltaResult,
    file_a_name: str,
    file_b_name: str,
    run_timestamp: str,
) -> list[tuple[str, str]]:
    """
    Return a list of (label, text) pairs for the Executive Narrative tab.
    All numbers are filled in from the actual DeltaResult.
    """
    total_a   = result.total_a
    total_b   = result.total_b
    n_only_a  = len(result.only_in_a)
    n_only_b  = len(result.only_in_b)
    n_matched = len(result.matched)
    n_changed = len(result.changed)
    n_dup_a   = len(result.duplicates_a)
    n_dup_b   = len(result.duplicates_b)
    n_blank_a = len(result.blank_keys_a)
    n_blank_b = len(result.blank_keys_b)

    def pct(n, d):
        return f"{n / d * 100:.1f}%" if d else "N/A"

    key_label = " + ".join(result.key_cols_a) if result.key_cols_a else "configured match key(s)"

    overview = (
        f"A bidirectional reconciliation was performed between the source dataset ({file_a_name}, "
        f"{total_a:,} records) and the comparison dataset ({file_b_name}, {total_b:,} records). "
        f"Records were matched using the identifier field(s): {key_label}. "
        f"Analysis completed: {run_timestamp}."
    )

    matching = (
        f"Of the {total_a:,} source records, {n_matched:,} ({pct(n_matched, total_a)}) were "
        f"matched to a corresponding record in the comparison dataset. "
        f"{n_only_a:,} records ({pct(n_only_a, total_a)} of source) appear exclusively in "
        f"{file_a_name} and are absent from the comparison dataset; these may represent "
        f"withdrawn entries, pending submissions, or records not yet reflected in the comparison dataset. "
        f"{n_only_b:,} records ({pct(n_only_b, total_b)} of comparison) appear exclusively in "
        f"{file_b_name} and have no corresponding source entry; these may represent new "
        f"submissions, late arrivals, or records not yet posted to the source dataset."
    )

    if result.compare_cols_a:
        compare_label = ", ".join(result.compare_cols_a)
        if n_changed == 0:
            changes = (
                f"All {n_matched:,} matched records were identical across the compared fields "
                f"({compare_label}). No field-level differences were detected."
            )
        else:
            changes = (
                f"Of the {n_matched:,} matched records, {n_changed:,} ({pct(n_changed, n_matched)}) "
                f"contained at least one field-level difference in the following compared fields: "
                f"{compare_label}. Before and after values for each difference are detailed in the "
                f"'Records with Differences' worksheet."
            )
    else:
        changes = (
            "Field-level comparison was not configured for this run. To identify value-level "
            "differences between matched records, re-run the analysis with one or more "
            "comparison fields selected."
        )

    if n_dup_a > 0 or n_dup_b > 0:
        duplicates = (
            f"{n_dup_a:,} source record(s) and {n_dup_b:,} comparison record(s) share a "
            f"match key with at least one other row in the same dataset. All duplicate rows are "
            f"documented in the identifier worksheets; only the first occurrence of each "
            f"identifier was used in the matching process. Duplicate submissions should be "
            f"investigated and resolved in the source system prior to any official reporting."
        )
    else:
        duplicates = (
            f"No duplicate identifiers were detected in either dataset. "
            f"Both {file_a_name} and {file_b_name} contain unique records per match key."
        )

    if n_blank_a > 0 or n_blank_b > 0:
        blanks = (
            f"{n_blank_a:,} source record(s) and {n_blank_b:,} comparison record(s) contained "
            f"a blank or null value in the match key field(s) and were excluded from the "
            f"reconciliation. These records are captured in the Data Quality Flags worksheet. "
            f"Source system corrections are required before these records can be included in "
            f"future analysis cycles."
        )
    else:
        blanks = (
            f"No missing identifier values were detected. All records in both datasets carried "
            f"valid, non-null match key values and were eligible for reconciliation."
        )

    recommendation = (
        f"Priority actions: "
        f"(1) Investigate {n_only_a:,} source-only record(s): confirm whether these represent "
        f"intentional removals, pending resubmissions, or data feed gaps. "
        f"(2) Validate {n_only_b:,} comparison-only record(s) against the authoritative source "
        f"of record before accepting as new entries. "
        f"(3) Review {n_changed:,} changed record(s) to determine whether "
        f"field-level differences reflect authorized updates or require correction. "
        f"(4) Resolve all data quality flags (duplicate identifiers and missing key values) "
        f"at the source system level."
    )

    return [
        ("Analysis Overview",       overview),
        ("Matching Results",        matching),
        ("Field-Level Differences", changes),
        ("Duplicate Identifiers",   duplicates),
        ("Missing Identifiers",     blanks),
        ("Recommended Actions",     recommendation),
    ]


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _build_delta_counts_df(result: DeltaResult) -> pd.DataFrame:
    """Flat reconciliation count table suitable for pivot tables and downstream reporting."""
    return pd.DataFrame([
        {"Category": "Source Dataset: Total Records",      "Count": result.total_a,              "Dataset": "Source"},
        {"Category": "Comparison Dataset: Total Records",  "Count": result.total_b,              "Dataset": "Comparison"},
        {"Category": "Source Only Records",                "Count": len(result.only_in_a),       "Dataset": "Source"},
        {"Category": "Comparison Only Records",            "Count": len(result.only_in_b),       "Dataset": "Comparison"},
        {"Category": "Matched Records",                    "Count": len(result.matched),         "Dataset": "Both"},
        {"Category": "Changed Records",                    "Count": len(result.changed),         "Dataset": "Both"},
        {"Category": "Source Duplicate Keys",              "Count": len(result.duplicates_a),    "Dataset": "Source"},
        {"Category": "Comparison Duplicate Keys",          "Count": len(result.duplicates_b),    "Dataset": "Comparison"},
        {"Category": "Source Missing Keys",                "Count": len(result.blank_keys_a),    "Dataset": "Source"},
        {"Category": "Comparison Missing Keys",            "Count": len(result.blank_keys_b),    "Dataset": "Comparison"},
    ])


def _write_sheet(df: pd.DataFrame, writer: pd.ExcelWriter, sheet_name: str) -> None:
    if df is not None and not df.empty:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    else:
        pd.DataFrame([["No records in this category."]]).to_excel(
            writer, sheet_name=sheet_name, index=False, header=False
        )


def _build_metadata_df(
    result: DeltaResult,
    file_a_name: str,
    file_b_name: str,
    run_timestamp: str,
    prep_meta_a: Optional[dict] = None,
    prep_meta_b: Optional[dict] = None,
) -> pd.DataFrame:
    """Tabular run metadata for the Analysis Metadata tab."""
    rows = [
        ("Analysis Timestamp",                   run_timestamp),
        ("Source Dataset",                       file_a_name),
        ("Source Sheet / Tab",                   result.sheet_a or "Default (first sheet or CSV)"),
        ("Source Record Count",                  str(result.total_a)),
        ("Comparison Dataset",                   file_b_name),
        ("Comparison Sheet / Tab",               result.sheet_b or "Default (first sheet or CSV)"),
        ("Comparison Record Count",              str(result.total_b)),
        ("Match Key Fields (Source)",            ", ".join(result.key_cols_a) if result.key_cols_a else "None"),
        ("Match Key Fields (Comparison)",        ", ".join(result.key_cols_b) if result.key_cols_b else "None"),
        ("Comparison Fields (Source)",           ", ".join(result.compare_cols_a) if result.compare_cols_a else "None configured"),
        ("Comparison Fields (Comparison)",       ", ".join(result.compare_cols_b) if result.compare_cols_b else "None configured"),
        ("Comparison Rules Applied",             str(len(result.comparison_rules))),
        ("Parse Warnings",                       str(len(result.compare_parse_issues)) if result.compare_parse_issues is not None else "0"),
    ]

    if prep_meta_a:
        rows.extend([
            ("Source Header Row Selected",        str(prep_meta_a.get("header_row_selected", ""))),
            ("Source Rows Dropped Above Header",  str(prep_meta_a.get("rows_dropped_above_header", ""))),
            ("Source Blank Rows Dropped",         str(prep_meta_a.get("rows_dropped_blank", ""))),
        ])
        if prep_meta_a.get("end_row_applied") is not None:
            rows.append(("Source End Row Applied", str(prep_meta_a["end_row_applied"])))

    if prep_meta_b:
        rows.extend([
            ("Comparison Header Row Selected",        str(prep_meta_b.get("header_row_selected", ""))),
            ("Comparison Rows Dropped Above Header",  str(prep_meta_b.get("rows_dropped_above_header", ""))),
            ("Comparison Blank Rows Dropped",         str(prep_meta_b.get("rows_dropped_blank", ""))),
        ])
        if prep_meta_b.get("end_row_applied") is not None:
            rows.append(("Comparison End Row Applied", str(prep_meta_b["end_row_applied"])))

    return pd.DataFrame(rows, columns=["Parameter", "Value"])


def _build_rules_df(result: DeltaResult) -> pd.DataFrame:
    """Tabular view of every comparison rule."""
    if not result.comparison_rules:
        return pd.DataFrame(
            [["No comparison rules configured."]],
            columns=["Note"],
        )
    rows = []
    for rule in result.comparison_rules:
        rows.append({
            "Source Field":     rule.get("column_a", ""),
            "Comparison Field": rule.get("column_b", ""),
            "Type":             rule.get("type", "text"),
            "Tolerance":        rule.get("tolerance", "") if rule.get("tolerance") is not None else "",
            "Date Mode":        rule.get("date_mode", "") if rule.get("date_mode") is not None else "",
        })
    return pd.DataFrame(rows)


def _write_narrative(
    narrative: list[tuple[str, str]],
    writer: pd.ExcelWriter,
) -> None:
    """Write the Executive Summary as a two-column label/text sheet."""
    rows = [{"Section": label, "Narrative": text} for label, text in narrative]
    pd.DataFrame(rows).to_excel(writer, sheet_name="Executive Summary", index=False)


def _apply_styles(wb) -> None:
    """Apply header colors, row alternation, column widths, and frozen panes."""
    white_bold = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        bottom=Side(style="thin", color="CCCCCC")
    )

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        fill_hex = _HEADER_COLORS.get(sheet_name, COLORS["navy"])
        header_fill = PatternFill(fill_type="solid", fgColor=fill_hex)
        alt_fill = PatternFill(fill_type="solid", fgColor=COLORS["light_gray"])
        row_highlight = _ROW_FILL_COLORS.get(sheet_name)
        highlight_fill = PatternFill(fill_type="solid", fgColor=row_highlight) if row_highlight else None

        # Header row
        for cell in ws[1]:
            cell.font = white_bold
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 22

        # Data rows
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            use_fill = highlight_fill if highlight_fill else (alt_fill if row_idx % 2 == 0 else None)
            for cell in row:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                cell.border = thin_border
                if use_fill:
                    cell.fill = use_fill

        # Narrative tab: wider text column, taller rows
        if sheet_name == "Executive Summary":
            ws.column_dimensions["A"].width = 28
            ws.column_dimensions["B"].width = 100
            for row_idx in range(2, ws.max_row + 1):
                ws.row_dimensions[row_idx].height = 80
                for cell in ws[row_idx]:
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        else:
            # Auto-fit all other columns (capped at 60 chars)
            for col_idx, col_cells in enumerate(ws.columns, start=1):
                max_len = 0
                for cell in col_cells:
                    try:
                        cell_len = len(str(cell.value)) if cell.value is not None else 0
                        max_len = max(max_len, cell_len)
                    except Exception:
                        pass
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)

        ws.freeze_panes = ws["A2"]
